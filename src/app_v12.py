"""
Fischer Energy Partners - Data Processing Application V12
Multi-Column CSV Support with Automatic Workflow

NEW IN V12:
- Multi-column CSV support: AI detects and allows selection of multiple value columns
- Checkbox interface for CSV column selection (matches Excel workflow)
- Column naming: "{Filename} {ColumnName}" prefix for uniqueness
- Backward compatible: Single-column CSVs still work

All V11 Features:
- Single-button automatic processing (combine -> save CSV -> resample -> generate Excel)
- Real-time progress bar with phase-weighted tracking (0-100%)
- Automatic file generation to archive folder with timestamped filenames
- Streamlined 4-step process
- Tab-based UI, smart data types, enhanced flagging, multi-tab Excel
"""

import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
import warnings
import json
import os
import re
import io
import base64
from dotenv import load_dotenv
from anthropic import Anthropic
from concurrent.futures import ThreadPoolExecutor, as_completed
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import shutil

# Add src directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))
from timestamp_normalizer import format_timestamp_mdy_hms, detect_timestamp_format

warnings.filterwarnings('ignore')

# Load environment variables
load_dotenv()

# Page configuration
st.set_page_config(
    page_title="Fischer Data Processing App",
    page_icon="📊",
    layout="wide"
)

# Initialize session state
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = {}
if 'file_configs' not in st.session_state:
    st.session_state.file_configs = {}
if 'combined_df' not in st.session_state:
    st.session_state.combined_df = None
if 'resampled_df' not in st.session_state:
    st.session_state.resampled_df = None
if 'resampling_stats' not in st.session_state:
    st.session_state.resampling_stats = {}
if 'inexact_cells' not in st.session_state:
    st.session_state.inexact_cells = pd.DataFrame()  # Boolean DataFrame for inexact cell tracking (memory-efficient)
if 'ai_debug_log' not in st.session_state:
    st.session_state.ai_debug_log = []
if 'ai_analysis_complete' not in st.session_state:
    st.session_state.ai_analysis_complete = False
if 'building_name' not in st.session_state:
    st.session_state.building_name = ""
if 'archive_path' not in st.session_state:
    st.session_state.archive_path = ""
if 'raw_csv_path' not in st.session_state:
    st.session_state.raw_csv_path = None
if 'excel_output_path' not in st.session_state:
    st.session_state.excel_output_path = None
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False


def add_logo():
    """Add Fischer Energy logo to top-left with title."""
    logo_path = Path(__file__).parent.parent / "assets" / "fischer background clear (1).png"

    if logo_path.exists():
        with open(logo_path, "rb") as f:
            logo_data = base64.b64encode(f.read()).decode()

        st.markdown(
            f"""
            <div style="display: flex; align-items: center; padding: 10px 0; margin-bottom: 20px;">
                <img src="data:image/png;base64,{logo_data}"
                     style="height: 80px; margin-right: 20px;">
                <h1 style="color: #24b3aa; margin: 0; font-size: 2.5rem;">Fischer Data Processor V12</h1>
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.title("🔧 Fischer Data Processing App")


def inject_custom_css():
    """Inject custom CSS for Fischer Energy branding."""
    st.markdown(
        """
        <style>
        /* Button styling */
        .stButton > button {
            background-color: #24b3aa;
            color: #FFFFFF;
            border: none;
            border-radius: 4px;
            padding: 0.5rem 1rem;
            font-weight: 500;
        }
        .stButton > button:hover {
            background-color: #1e9b93;
        }

        /* Primary button styling */
        .stButton > button[kind="primary"] {
            background-color: #24b3aa;
            color: #FFFFFF;
        }
        .stButton > button[kind="primary"]:hover {
            background-color: #1e9b93;
        }

        /* Header styling */
        h1, h2, h3 {
            color: #24b3aa;
        }

        /* Expander headers */
        .streamlit-expanderHeader {
            background-color: #f0f9f8;
            border-left: 3px solid #24b3aa;
            font-weight: 500;
        }

        /* Success messages */
        .stSuccess {
            background-color: #e6f7f6;
            border-left: 4px solid #24b3aa;
        }

        /* Info messages */
        .stInfo {
            background-color: #f0f9f8;
            border-left: 4px solid #24b3aa;
        }

        /* Dataframe headers */
        .dataframe thead tr th {
            background-color: #24b3aa !important;
            color: #FFFFFF !important;
        }

        /* Sidebar styling */
        section[data-testid="stSidebar"] {
            background-color: #f0f9f8;
        }

        /* Download button */
        .stDownloadButton > button {
            background-color: #24b3aa;
            color: #FFFFFF;
        }
        .stDownloadButton > button:hover {
            background-color: #1e9b93;
        }

        /* Tab container styling - adds teal border around entire tab interface */
        .stTabs {
            border: 2px solid #24b3aa;
            border-radius: 8px;
            padding: 16px;
            background-color: #ffffff;
            margin: 16px 0;
        }

        /* Tab list styling - separator between tabs */
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px;
            border-bottom: 2px solid #e0e0e0;
            padding-bottom: 8px;
            margin-bottom: 16px;
        }

        /* Individual tab styling */
        .stTabs [data-baseweb="tab"] {
            border-right: 1px solid #e0e0e0;
            padding-right: 16px;
            padding-left: 16px;
        }

        /* Remove border from last tab */
        .stTabs [data-baseweb="tab"]:last-child {
            border-right: none;
        }

        /* Active tab styling */
        .stTabs [aria-selected="true"] {
            border-bottom: 3px solid #24b3aa;
            color: #24b3aa;
            font-weight: 600;
        }

        /* Tab panel content styling */
        .stTabs [data-baseweb="tab-panel"] {
            padding: 16px;
            background-color: #fafafa;
            border-radius: 4px;
            margin-top: 8px;
        }

        /* Nested tabs styling (for sheet-level tabs inside file-level tabs) */
        .stTabs .stTabs {
            border: 1px solid #d0d0d0;
            border-radius: 4px;
            padding: 12px;
            background-color: #ffffff;
            margin-top: 8px;
        }

        .stTabs .stTabs [data-baseweb="tab-list"] {
            border-bottom: 1px solid #d0d0d0;
        }
        </style>
        """,
        unsafe_allow_html=True
    )


# Apply branding
add_logo()
inject_custom_css()


def read_raw_lines(file_path, num_lines=15):
    """Read first N lines of a file as raw text. For Excel files, convert to CSV-like format."""
    lines = []

    # Check if it's an Excel file
    if str(file_path).lower().endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(file_path, header=None, nrows=num_lines,
                             dtype=str, keep_default_na=False)
            csv_buffer = io.StringIO()
            df.to_csv(csv_buffer, index=False, header=False)
            csv_text = csv_buffer.getvalue()
            lines = csv_text.strip().split('\n')
            return lines[:num_lines]
        except Exception as e:
            return [f"Error reading Excel file: {str(e)}"]

    # For CSV/text files
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f):
                if i >= num_lines:
                    break
                lines.append(line.rstrip())
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='latin-1') as f:
            for i, line in enumerate(f):
                if i >= num_lines:
                    break
                lines.append(line.rstrip())
    return lines


def detect_file_type(file_path):
    """
    Detect if file is CSV or Excel, and if Excel has multiple tabs.

    Returns:
        Tuple of (file_type, sheet_names)
        - file_type: 'excel_multi_tab', 'excel_single_tab', or 'csv'
        - sheet_names: List of tab names (None for CSV files)
    """
    if str(file_path).lower().endswith(('.xlsx', '.xls')):
        try:
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
            num_tabs = len(sheet_names)

            if num_tabs > 1:
                return 'excel_multi_tab', sheet_names
            else:
                return 'excel_single_tab', sheet_names
        except Exception as e:
            return 'csv', None
    else:
        return 'csv', None


def read_tab_raw_lines(file_path, sheet_name, num_lines=15):
    """Read first N lines from a specific Excel tab."""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=num_lines,
                         dtype=str, keep_default_na=False)
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, header=False)
        csv_text = csv_buffer.getvalue()
        lines = csv_text.strip().split('\n')
        return lines[:num_lines]
    except Exception as e:
        return [f"Error reading tab {sheet_name}: {str(e)}"]


def build_ai_prompt(file_name, text_sample):
    """
    Build the AI prompt for column detection.

    V12: Now asks for MULTIPLE value columns (array) instead of single value_column.
    V12+: Also detects stacked/long format with equipment identifier columns and split date/time.
    """
    return f"""Analyze the first 15 lines of this CSV/Excel file sample to determine the column configuration for data processing.

File name: {file_name or 'unknown'}

Raw file content (first 15 lines):
{text_sample}

Based on this data, identify:
1. The delimiter used (comma, tab, semicolon, etc.)
2. Which row index (0-based) contains the column headers and is the START of data
3. Which column index (0-based) contains the date/timestamp
4. Which column indices (0-based) contain sensor values/readings - there may be MULTIPLE value columns
5. The names of the value columns (extract from headers or suggest descriptive names)
6. Whether the file uses a "stacked/long" format where multiple equipment/sensors share the same timestamp rows and an identifier column (like "Equipment Name") distinguishes which equipment each row belongs to. In stacked format, data for one equipment is listed first (e.g. all timestamps for HX-01), then repeated for the next equipment (HX-02), etc.
7. Whether the date and time are in SEPARATE columns (e.g. "Property Date" in one column, "Property Time" in another)

Return ONLY a JSON object in this exact format with no additional text:
{{
  "delimiter": ",",
  "start_row": 1,
  "date_column": 0,
  "value_columns": [2, 3, 4],
  "column_names": ["Temperature", "Humidity", "Pressure"],
  "is_stacked": false,
  "equipment_column": null,
  "time_column": null
}}

Rules:
- All indices MUST be 0-based
- start_row is where column headers are located (data begins on start_row+1)
- value_columns should be a LIST of column indices (can be one or multiple)
- column_names should match the order of value_columns
- is_stacked: set to true if data has an equipment/identifier column and rows for different equipment share the same timestamps
- equipment_column: 0-based column index of the equipment/identifier column (null if not stacked)
- time_column: 0-based column index of a SEPARATE time column if date and time are split across two columns (null if timestamp is in a single column). Do NOT include the time column in value_columns.
- If a column doesn't exist, use -1 as the value
- Return only valid JSON with no explanations or additional text"""


def build_multi_tab_ai_prompt(file_name, tabs_data):
    """
    Build AI prompt for multi-tab Excel file analysis.

    Args:
        file_name: Name of the Excel file
        tabs_data: Dict of {tab_name: text_sample} for each tab

    Returns:
        Formatted prompt string for AI analysis
    """
    tabs_text = ""
    for tab_name, text_sample in tabs_data.items():
        tabs_text += f"\n\n=== TAB: {tab_name} ===\n{text_sample}\n"

    return f"""Analyze this multi-tab Excel file to determine column configurations for data processing.

File name: {file_name or 'unknown'}

{tabs_text}

For EACH tab, identify:
1. Which row index (0-based) contains the column headers (start_row)
2. Which column index (0-based) contains the date/timestamp
3. Which column indices (0-based) contain sensor values/readings - MULTIPLE columns are expected per tab
4. The names of the value columns (extract from headers or suggest descriptive names)

Return ONLY a JSON object in this exact format with no additional text:
{{
  "tabs": [
    {{
      "tab_name": "AC12-1",
      "start_row": 1,
      "date_column": 0,
      "value_columns": [2, 3, 4],
      "column_names": ["Return Air Temp", "Supply Air Temp", "Fan Status"]
    }},
    {{
      "tab_name": "AC12-2",
      "start_row": 1,
      "date_column": 0,
      "value_columns": [2, 3],
      "column_names": ["Return Air Temp", "Supply Air Temp"]
    }}
  ]
}}

Rules:
- All indices MUST be 0-based
- Analyze ALL tabs provided above
- value_columns should be a LIST of column indices (can be multiple per tab)
- column_names should match the order of value_columns
- Return only valid JSON with no explanations"""


def call_claude_api(prompt, api_key):
    """Call Claude API and return the response text."""
    client = Anthropic(api_key=api_key)

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=15000,
            temperature=0.6,
            messages=[{"role": "user", "content": prompt}]
        )

        # Extract text from response
        response_text = response.content[0].text.strip()
        return response_text
    except Exception as e:
        raise RuntimeError(f"Claude API call failed: {str(e)}")


def analyze_single_file(file_name, file_path, api_key):
    """
    Analyze a single CSV or single-tab Excel file with AI.

    V12: Now handles multi-column response format and converts to internal structure.

    Returns tuple of (config, debug_entry)
    Config format:
    {
        'file_type': 'csv',
        'config': {
            'start_row': 1,
            'delimiter': ',',
            'date_column': 0,
            'available_columns': [2, 3, 4],
            'column_names': ['Temp', 'Humidity', 'Pressure'],
            'selected_columns': [2, 3, 4]  # Initially all selected
        }
    }
    """
    debug_entry = {
        'file_name': file_name,
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'request': None,
        'response': None,
        'error': None,
        'success': False
    }

    try:
        # Read first 15 lines
        raw_lines = read_raw_lines(file_path, num_lines=15)
        text_sample = "\n".join(raw_lines)

        # Build prompt
        prompt = build_ai_prompt(file_name, text_sample)

        # Store request
        debug_entry['request'] = {
            'model': 'claude-sonnet-4-5-20250929',
            'max_tokens': 15000,
            'temperature': 0.6,
            'prompt_length': len(prompt),
            'prompt_preview': prompt[:500] + '...' if len(prompt) > 500 else prompt,
            'raw_text_lines': raw_lines[:5]
        }

        # Call API
        response_text = call_claude_api(prompt, api_key)

        # Store response
        debug_entry['response'] = {
            'raw_text': response_text,
            'response_length': len(response_text)
        }

        # Parse JSON
        json_start = response_text.find('{')
        json_end = response_text.rfind('}') + 1
        if json_start != -1 and json_end > json_start:
            json_str = response_text[json_start:json_end]
            parsed = json.loads(json_str)

            # V12: Convert to internal format with multi-column support
            # Handle backward compatibility with old single-column format
            if 'value_column' in parsed and 'value_columns' not in parsed:
                # Old format - convert to new array format
                value_cols = [parsed['value_column']]
                col_names = [parsed.get('sensor_name', f"Column_{parsed['value_column']}")]
            else:
                # New format - use arrays
                value_cols = parsed.get('value_columns', [2])
                col_names = parsed.get('column_names', [f"Column_{i}" for i in value_cols])

            # Detect stacked/long format fields
            is_stacked = parsed.get('is_stacked', False)
            equipment_column = parsed.get('equipment_column', None)
            time_column = parsed.get('time_column', None)

            # Build internal config structure (matches Excel multi-tab format)
            config = {
                'start_row': parsed.get('start_row', 0),
                'delimiter': parsed.get('delimiter', ','),
                'date_column': parsed.get('date_column', 0),
                'available_columns': value_cols,
                'column_names': col_names,
                'selected_columns': value_cols.copy()  # Initially all selected
            }

            # Add stacked-specific fields if detected
            if is_stacked and equipment_column is not None:
                config['is_stacked'] = True
                config['equipment_column'] = equipment_column
                config['time_column'] = time_column

            debug_entry['response']['parsed_json'] = config
            debug_entry['success'] = True

            return config, debug_entry
        else:
            debug_entry['error'] = "Could not find JSON in response"
            return None, debug_entry

    except json.JSONDecodeError as e:
        debug_entry['error'] = f"JSON decode error: {str(e)}"
        return None, debug_entry
    except Exception as e:
        debug_entry['error'] = f"Error: {str(e)}"
        return None, debug_entry


def analyze_multi_tab_file(file_name, file_path, sheet_names, api_key):
    """
    Analyze a multi-tab Excel file with AI.

    Returns tuple of (config, debug_entry)
    Config format: {
        "file_type": "excel_multi_tab",
        "tabs": {
            "Tab1": {
                "start_row": 1,
                "date_column": 0,
                "available_columns": [2, 3, 4],
                "column_names": ["Col1", "Col2", "Col3"],
                "selected_columns": [2, 3, 4]  # Initially all selected
            }
        }
    }
    """
    debug_entry = {
        'file_name': file_name,
        'timestamp': datetime.now().strftime('%H:%M:%S'),
        'request': None,
        'response': None,
        'error': None,
        'success': False
    }

    try:
        # Read first 15 lines from each tab
        tabs_data = {}
        for sheet_name in sheet_names:
            raw_lines = read_tab_raw_lines(file_path, sheet_name, num_lines=15)
            tabs_data[sheet_name] = "\n".join(raw_lines)

        # Build multi-tab prompt
        prompt = build_multi_tab_ai_prompt(file_name, tabs_data)

        # Store request
        debug_entry['request'] = {
            'model': 'claude-sonnet-4-5-20250929',
            'max_tokens': 15000,  # Increased for multiple tabs
            'temperature': 0.6,
            'prompt_length': len(prompt),
            'prompt_preview': prompt[:500] + '...' if len(prompt) > 500 else prompt,
            'tabs_analyzed': list(sheet_names)
        }

        # Call API
        response_text = call_claude_api(prompt, api_key)

        # Store response
        debug_entry['response'] = {
            'raw_text': response_text,
            'response_length': len(response_text)
        }

        # Parse JSON
        json_start = response_text.find('{')
        json_end = response_text.rfind('}') + 1
        if json_start != -1 and json_end > json_start:
            json_str = response_text[json_start:json_end]
            ai_response = json.loads(json_str)

            # Convert AI response to our internal format
            config = {
                "file_type": "excel_multi_tab",
                "tabs": {}
            }

            for tab_data in ai_response.get('tabs', []):
                tab_name = tab_data['tab_name']
                value_cols = tab_data['value_columns']

                config['tabs'][tab_name] = {
                    'start_row': tab_data['start_row'],
                    'date_column': tab_data['date_column'],
                    'available_columns': value_cols,
                    'column_names': tab_data['column_names'],
                    'selected_columns': value_cols.copy()  # Initially all selected
                }

            debug_entry['response']['parsed_json'] = config
            debug_entry['success'] = True

            return config, debug_entry
        else:
            debug_entry['error'] = "Could not find JSON in response"
            return None, debug_entry

    except json.JSONDecodeError as e:
        debug_entry['error'] = f"JSON decode error: {str(e)}"
        return None, debug_entry
    except Exception as e:
        debug_entry['error'] = f"Error: {str(e)}"
        return None, debug_entry


def analyze_file_with_detection(file_name, file_path, api_key):
    """
    Detect file type and analyze accordingly.

    Returns tuple of (config, debug_entry)
    """
    # Detect file type
    file_type, sheet_names = detect_file_type(file_path)

    if file_type == 'excel_multi_tab':
        # Analyze multi-tab Excel file
        config, debug_entry = analyze_multi_tab_file(file_name, file_path, sheet_names, api_key)
        if config:
            config['file_type'] = 'excel_multi_tab'
    else:
        # Analyze CSV or single-tab Excel file
        config, debug_entry = analyze_single_file(file_name, file_path, api_key)
        if config:
            # Check if AI detected stacked/long format
            if config.get('is_stacked', False) and config.get('equipment_column') is not None:
                config = {
                    'file_type': 'stacked_long',
                    'config': config
                }
            else:
                # Wrap in standard format for consistency
                config = {
                    'file_type': 'csv' if file_type == 'csv' else 'excel_single_tab',
                    'config': config
                }

    return config, debug_entry


def analyze_all_files_parallel(uploaded_files, api_key):
    """
    Analyze all uploaded files in parallel using ThreadPoolExecutor.

    V9: Now detects and handles multi-tab Excel files automatically.
    """
    configs = {}
    debug_logs = []

    with ThreadPoolExecutor(max_workers=5) as executor:
        # Submit all tasks (now with file type detection)
        future_to_file = {
            executor.submit(analyze_file_with_detection, file_name, file_path, api_key): file_name
            for file_name, file_path in uploaded_files.items()
        }

        # Collect results as they complete
        for future in as_completed(future_to_file):
            file_name = future_to_file[future]
            try:
                config, debug_entry = future.result()
                if config:
                    configs[file_name] = config
                debug_logs.append(debug_entry)
            except Exception as e:
                debug_logs.append({
                    'file_name': file_name,
                    'timestamp': datetime.now().strftime('%H:%M:%S'),
                    'error': f"Exception: {str(e)}",
                    'success': False
                })

    return configs, debug_logs


def parse_file_with_config(file_path, start_row=0, delimiter=',', num_rows=10):
    """Parse file using the provided configuration."""
    try:
        if str(file_path).lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_path, header=start_row, nrows=num_rows,
                             dtype=str, keep_default_na=False)
        else:
            df = pd.read_csv(
                file_path,
                sep=delimiter,
                header=start_row,
                nrows=num_rows,
                dtype=str,
                keep_default_na=False,
                encoding='utf-8',
                encoding_errors='ignore',
                on_bad_lines='skip'
            )
        return df
    except Exception as e:
        return None


def calculate_zero_flags(resampled_df, sensor_cols):
    """
    Calculate Zero_Value_Flag for each row based on zero patterns across all sensors.

    Logic:
    - "Clear": No zeros detected in any sensor for this row
    - "Single": At least one sensor has a zero that is NOT consecutive with previous row
    - "Repeated": At least one sensor has 2+ consecutive zeros (current and previous both zero)

    Priority: "Repeated" > "Single" > "Clear"

    V10: Only checks numeric columns (skips text fields like "off"/"on")

    Args:
        resampled_df: The resampled DataFrame
        sensor_cols: List of sensor column names

    Returns:
        List of zero flag values for each row
    """
    zero_flags = []

    for idx in range(len(resampled_df)):
        row_flag = "Clear"

        for sensor in sensor_cols:
            # Skip text columns (only check numeric columns for zeros)
            if not pd.api.types.is_numeric_dtype(resampled_df[sensor]):
                continue

            val_current = resampled_df.loc[idx, sensor]

            # Skip if current value is NaN or not zero
            if pd.isna(val_current) or val_current != 0:
                continue

            # Current value is zero
            if idx > 0:
                val_prev = resampled_df.loc[idx-1, sensor]
                # Check if previous is also zero (repeated)
                if val_prev == 0:
                    row_flag = "Repeated"
                    break  # Highest priority, stop checking
                else:
                    # Single zero (not consecutive)
                    if row_flag == "Clear":
                        row_flag = "Single"
            else:
                # First row, can't be repeated
                if row_flag == "Clear":
                    row_flag = "Single"

        zero_flags.append(row_flag)

    return zero_flags


def resample_to_quarter_hour(combined_df, tolerance_minutes=2, progress_callback=None):
    """
    Resample combined data to 15-minute intervals with PER-SENSOR nearest-value matching.

    V11 Optimized: Uses pd.merge_asof for O(n log n) performance instead of O(n*m).
    This version is memory-efficient and works within Replit's resource limits.

    V9 Features (preserved):
    - Modified stale flag: 3+ consecutive non-zero identical values (changed from 4+)
    - New Zero_Value_Flag: Clear/Single/Repeated per-sensor tracking
    - Each sensor independently finds its closest value within +/-2 minutes
    - Tracks inexact cells for Excel color-coding
    - Consolidated stale flags: single True/False column + comma-separated sensor list

    Args:
        combined_df: The merged dataframe with all sensors
        tolerance_minutes: Window for finding nearest match (+/-2 minutes default)
        progress_callback: Optional callback for progress updates

    Returns:
        Tuple of (resampled_df, stats_dict, inexact_df) where inexact_df is a Boolean DataFrame
    """
    if combined_df is None or combined_df.empty:
        return None, {}, pd.DataFrame()

    try:
        # Create complete range of 15-minute timestamps
        start_time = combined_df['Date'].min()
        end_time = combined_df['Date'].max()

        # Round start to previous 15-min mark
        start_time = start_time.replace(minute=(start_time.minute // 15) * 15, second=0, microsecond=0)

        # Round end to next 15-min mark
        end_minute = ((end_time.minute // 15) + 1) * 15
        if end_minute >= 60:
            end_time = end_time + timedelta(hours=1)
            end_time = end_time.replace(minute=0, second=0, microsecond=0)
        else:
            end_time = end_time.replace(minute=end_minute, second=0, microsecond=0)

        # Generate 15-minute interval timestamps
        target_timestamps = pd.date_range(start=start_time, end=end_time, freq='15min')

        # Get sensor columns (exclude Date)
        sensor_cols = [col for col in combined_df.columns if col != 'Date']

        # Log data size for debugging
        num_intervals = len(target_timestamps)
        num_sensors = len(sensor_cols)
        num_rows = len(combined_df)

        if progress_callback:
            progress_callback(0, num_sensors,
                f"Preparing resampling: {num_intervals} intervals, {num_sensors} sensors, {num_rows} source rows")

        # Create target DataFrame with just dates
        target_df = pd.DataFrame({'Date': target_timestamps})

        # Sort combined_df by Date once (required for merge_asof)
        combined_sorted = combined_df.sort_values('Date').reset_index(drop=True)

        # Tolerance for merge_asof
        tolerance = pd.Timedelta(minutes=tolerance_minutes)

        # Initialize result with target timestamps
        resampled = target_df.copy()

        # Track inexact cells using Boolean DataFrame (memory-efficient for wide datasets)
        # This replaces the nested dictionary which caused memory issues with 1000+ sensors
        inexact_df = pd.DataFrame(index=range(len(target_timestamps)))
        total_inexact = 0

        # Process each sensor using merge_asof (memory-efficient, O(n log n))
        for sensor_idx, sensor in enumerate(sensor_cols):
            if progress_callback and sensor_idx % 5 == 0:
                progress_callback(sensor_idx, num_sensors,
                    f"Resampling sensor {sensor_idx + 1}/{num_sensors}: {sensor}")

            # Extract only Date and this sensor's values (drop NaN to save memory)
            sensor_data = combined_sorted[['Date', sensor]].dropna(subset=[sensor]).copy()

            if sensor_data.empty:
                # No data for this sensor - fill with NaN
                resampled[sensor] = None
                inexact_df[sensor] = False  # Vectorized: entire column is False
                continue

            # Ensure sensor_data is sorted by Date
            sensor_data = sensor_data.sort_values('Date').reset_index(drop=True)

            # Use merge_asof to find nearest value within tolerance (forward direction)
            merged_forward = pd.merge_asof(
                target_df,
                sensor_data.rename(columns={sensor: f'{sensor}_fwd', 'Date': 'Date_fwd'}),
                left_on='Date',
                right_on='Date_fwd',
                direction='forward',
                tolerance=tolerance
            )

            # Use merge_asof to find nearest value within tolerance (backward direction)
            merged_backward = pd.merge_asof(
                target_df,
                sensor_data.rename(columns={sensor: f'{sensor}_bwd', 'Date': 'Date_bwd'}),
                left_on='Date',
                right_on='Date_bwd',
                direction='backward',
                tolerance=tolerance
            )

            # Vectorized selection of closer forward or backward match
            fwd_val = merged_forward[f'{sensor}_fwd']
            bwd_val = merged_backward[f'{sensor}_bwd']
            fwd_time = merged_forward['Date_fwd']
            bwd_time = merged_backward['Date_bwd']

            # Calculate time differences (vectorized)
            fwd_diff = (fwd_time - target_df['Date']).abs()
            bwd_diff = (bwd_time - target_df['Date']).abs()

            # Determine validity masks
            fwd_valid = fwd_val.notna()
            bwd_valid = bwd_val.notna()

            # Default to backward values
            sensor_values = bwd_val.copy()
            source_times = bwd_time.copy()

            # Where only forward is valid, use forward
            only_fwd = fwd_valid & ~bwd_valid
            sensor_values = sensor_values.where(~only_fwd, fwd_val)
            source_times = source_times.where(~only_fwd, fwd_time)

            # Where both are valid, pick the closer one (forward wins ties)
            both_valid = fwd_valid & bwd_valid
            use_fwd = both_valid & (fwd_diff <= bwd_diff)
            sensor_values = sensor_values.where(~use_fwd, fwd_val)
            source_times = source_times.where(~use_fwd, fwd_time)

            # Where neither is valid, set to None
            neither_valid = ~fwd_valid & ~bwd_valid
            sensor_values = sensor_values.where(~neither_valid, None)

            resampled[sensor] = sensor_values.values

            # Vectorized inexact cell tracking
            source_minute = source_times.dt.minute
            source_second = source_times.dt.second
            is_exact = ((source_minute % 15) == 0) & (source_second == 0)
            is_inexact = source_times.notna() & ~is_exact

            # Store as DataFrame column (vectorized - no loop needed)
            inexact_df[sensor] = is_inexact.values

            total_inexact += int(is_inexact.sum())

            # Free memory from temporary DataFrames
            del sensor_data, merged_forward, merged_backward

        if progress_callback:
            progress_callback(num_sensors, num_sensors, "Applying quality flags...")

    except MemoryError:
        rows_info = f"{num_rows} rows" if 'num_rows' in dir() else "data"
        sensors_info = f"{num_sensors} sensors" if 'num_sensors' in dir() else "sensors"
        raise MemoryError(
            f"Not enough memory to process {rows_info} with {sensors_info}. "
            "Try processing fewer files at once."
        )
    except Exception as e:
        raise RuntimeError(f"Resampling failed: {str(e)}")

    # Flag stale data per sensor (temporary for consolidation)
    # V9: Changed to 3+ consecutive non-zero values (was 4+ in V8)
    # V10: Only check numeric columns (skip text fields like "off"/"on")
    stale_per_sensor = {}
    for sensor in sensor_cols:
        # Skip text columns (check if column is numeric)
        if not pd.api.types.is_numeric_dtype(resampled[sensor]):
            # Text column - don't flag for staleness
            stale_per_sensor[sensor] = pd.Series([False] * len(resampled), index=resampled.index)
            continue

        # Skip if value is zero or NaN
        is_non_zero = (resampled[sensor] != 0) & (resampled[sensor].notna())

        # Check if current equals previous 2 values (3 consecutive identical non-zero)
        is_stale = (
            is_non_zero &
            (resampled[sensor] == resampled[sensor].shift(1)) &
            (resampled[sensor] == resampled[sensor].shift(2))
        )
        stale_per_sensor[sensor] = is_stale

    # Consolidate stale flags into two columns
    stale_data_flag = []
    stale_sensors_list = []

    for idx in range(len(resampled)):
        stale_sensors = [sensor for sensor in sensor_cols if stale_per_sensor[sensor].iloc[idx]]

        if stale_sensors:
            stale_data_flag.append(True)
            stale_sensors_list.append(', '.join(stale_sensors))
        else:
            stale_data_flag.append(False)
            stale_sensors_list.append('')

    # Add consolidated stale columns
    resampled['Stale_Data_Flag'] = stale_data_flag
    resampled['Stale_Sensors'] = stale_sensors_list

    # Calculate and add Zero_Value_Flag column (V9 new feature)
    zero_flags = calculate_zero_flags(resampled, sensor_cols)
    resampled['Zero_Value_Flag'] = zero_flags

    # Calculate statistics
    zero_flag_counts = {
        'Clear': zero_flags.count('Clear'),
        'Single': zero_flags.count('Single'),
        'Repeated': zero_flags.count('Repeated')
    }
    stale_counts = {sensor: int(stale_per_sensor[sensor].sum()) for sensor in sensor_cols}
    total_stale_flags = sum(stale_counts.values())

    stats = {
        'total_intervals': len(resampled),
        'total_inexact_cells': int(total_inexact),
        'stale_by_sensor': stale_counts,
        'total_stale_flags': total_stale_flags,
        'rows_with_stale_data': int(sum(stale_data_flag)),
        'zero_flag_counts': zero_flag_counts,
        'date_range': {
            'start': resampled['Date'].min(),
            'end': resampled['Date'].max()
        }
    }

    # Reorder columns: Date, flags, then sensor columns
    flag_cols = ['Stale_Data_Flag', 'Stale_Sensors', 'Zero_Value_Flag']
    sensor_cols_ordered = [col for col in resampled.columns if col not in ['Date'] + flag_cols]
    column_order = ['Date'] + flag_cols + sensor_cols_ordered
    resampled = resampled[column_order]

    return resampled, stats, inexact_df


def prepare_df_for_display(df):
    """
    Ensure DataFrame is PyArrow-compatible for st.dataframe() display.

    Fixes mixed data types that cause ArrowTypeError during serialization.

    Args:
        df: Pandas DataFrame to prepare

    Returns:
        Copy of DataFrame with corrected types
    """
    if df is None or df.empty:
        return df

    df_copy = df.copy()

    # Convert Date column to datetime if it's object type
    if 'Date' in df_copy.columns:
        if df_copy['Date'].dtype == 'object':
            df_copy['Date'] = pd.to_datetime(df_copy['Date'], errors='coerce')

    # Convert all other object columns to numeric where possible
    for col in df_copy.columns:
        if col not in ['Date', 'Stale_Sensors', 'Zero_Value_Flag', 'Stale_Data_Flag']:
            if df_copy[col].dtype == 'object':
                # Try to convert to numeric
                df_copy[col] = pd.to_numeric(df_copy[col], errors='ignore')

    return df_copy


def build_tab_label(base_name, selected_count, total_count):
    """Build tab label with visual indicators."""
    if selected_count > 0:
        return f"{base_name} ✓ ({selected_count})"
    elif total_count > 0:
        return f"{base_name} ⚠️"
    else:
        return f"{base_name} ❌"


def smart_convert_column(series, threshold=0.8):
    """
    Intelligently convert column to numeric or keep as text.

    Args:
        series: pandas Series to convert
        threshold: Minimum ratio of valid numeric values (default 0.8)

    Returns:
        Converted series (numeric) or original series (text)
    """
    # Try numeric conversion
    numeric_series = pd.to_numeric(series, errors='coerce')

    # Fast path: if ALL are NaN, it's pure text
    if numeric_series.isna().all():
        return series  # Keep original text

    # Check if mostly numeric
    valid_ratio = numeric_series.notna().sum() / len(numeric_series)

    if valid_ratio >= threshold:
        return numeric_series  # SQL-ready numeric
    else:
        return series  # Keep as text


def render_sheet_config_ui(file_name, file_path, sheet_name, config):
    """Render configuration UI for a single Excel sheet."""
    tab_config = config['tabs'][sheet_name]

    # Date column dropdown
    date_col_options = list(range(20))
    date_col_idx = tab_config.get('date_column', 0)

    date_column = st.selectbox(
        "Date Column",
        options=date_col_options,
        index=date_col_idx,
        key=f"date_{file_name}_{sheet_name}"
    )

    # Column selection checkboxes
    st.markdown("**Select Value Columns to Include:**")

    available_cols = tab_config.get('available_columns', [])
    column_names = tab_config.get('column_names', [])
    currently_selected = tab_config.get('selected_columns', available_cols.copy())

    new_selected = []
    for col_idx, col_name in zip(available_cols, column_names):
        is_checked = col_idx in currently_selected
        if st.checkbox(
            f"✓ **{col_name}** (column {col_idx})",
            value=is_checked,
            key=f"col_{file_name}_{sheet_name}_{col_idx}"
        ):
            new_selected.append(col_idx)

    # Update config
    tab_config['date_column'] = date_column
    tab_config['selected_columns'] = new_selected

    # Show preview
    if new_selected:
        st.success(f"✅ {len(new_selected)} column(s) selected from {sheet_name}")

        try:
            df_tab = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=tab_config['start_row'],
                nrows=5,
                dtype=str,
                keep_default_na=False
            )
            preview_cols = [df_tab.columns[date_column]] + \
                          [df_tab.columns[i] for i in new_selected if i < len(df_tab.columns)]
            st.dataframe(prepare_df_for_display(df_tab[preview_cols]), height=200)
        except Exception as e:
            st.caption(f"Preview unavailable: {str(e)}")
    else:
        st.warning(f"⚠️ No columns selected from {sheet_name}")


def render_csv_config_ui(file_name, file_path, config):
    """
    Render configuration UI for CSV or single-tab Excel file.

    V12: Now supports multi-column selection with checkbox interface (matches Excel workflow).
    """
    file_type = config.get('file_type', 'csv')
    inner_config = config.get('config', config)  # Handle both formats

    # Configuration inputs (row 1: basic settings)
    col1, col2, col3 = st.columns(3)

    with col1:
        start_row = st.number_input(
            "Start Row",
            min_value=0,
            max_value=10,
            value=inner_config.get('start_row', 0),
            key=f"start_{file_name}",
            help="Row where column headers are located"
        )

    with col2:
        delimiter_options = [',', '\t', ';', '|']
        delimiter_labels = ['Comma (,)', 'Tab (\\t)', 'Semicolon (;)', 'Pipe (|)']
        default_delim = inner_config.get('delimiter', ',')
        try:
            delim_idx = delimiter_options.index(default_delim)
        except ValueError:
            delim_idx = 0

        delimiter = st.selectbox(
            "Delimiter",
            options=delimiter_options,
            format_func=lambda x: delimiter_labels[delimiter_options.index(x)],
            index=delim_idx,
            key=f"delim_{file_name}"
        )

    with col3:
        date_column = st.number_input(
            "Date Column",
            min_value=0,
            max_value=20,
            value=inner_config.get('date_column', 0),
            key=f"date_{file_name}",
            help="Column index for date/timestamp (0-based)"
        )

    # V12: Multi-column selection with checkboxes
    st.markdown("**Select Value Columns to Include:**")

    available_cols = inner_config.get('available_columns', [])
    column_names = inner_config.get('column_names', [])
    currently_selected = inner_config.get('selected_columns', available_cols.copy())

    # If we have available columns, show checkboxes
    if available_cols:
        new_selected = []

        # Create columns for checkboxes (3 per row)
        checkbox_cols = st.columns(3)
        for i, (col_idx, col_name) in enumerate(zip(available_cols, column_names)):
            is_checked = col_idx in currently_selected
            with checkbox_cols[i % 3]:
                if st.checkbox(
                    f"✓ **{col_name}** (col {col_idx})",
                    value=is_checked,
                    key=f"col_{file_name}_{col_idx}"
                ):
                    new_selected.append(col_idx)

        # Show selected count
        if new_selected:
            st.success(f"✅ {len(new_selected)} column(s) selected")
        else:
            st.warning("⚠️ No columns selected")
    else:
        # Fallback for files without detected columns - allow manual column addition
        st.info("No value columns detected. Add columns manually below:")

        manual_col = st.number_input(
            "Add Value Column Index",
            min_value=0,
            max_value=20,
            value=2,
            key=f"manual_col_{file_name}",
            help="Column index for sensor values (0-based)"
        )

        manual_name = st.text_input(
            "Column Name",
            value=Path(file_name).stem,
            key=f"manual_name_{file_name}"
        )

        # Set manual values
        available_cols = [manual_col]
        column_names = [manual_name]
        new_selected = [manual_col]

    # Update config in session state with new multi-column format
    st.session_state.file_configs[file_name] = {
        'file_type': file_type,
        'config': {
            'start_row': start_row,
            'delimiter': delimiter,
            'date_column': date_column,
            'available_columns': available_cols,
            'column_names': column_names,
            'selected_columns': new_selected
        }
    }

    # Show extracted data preview
    st.markdown("**📊 Extracted Data Preview** (what will be used in processing):")
    df_preview = parse_file_with_config(file_path, start_row, delimiter, num_rows=10)

    if df_preview is not None and not df_preview.empty:
        # Create a preview of just the extracted columns
        extracted_preview = pd.DataFrame()

        # Get file prefix for column naming
        file_prefix = Path(file_name).stem

        # Show date column info
        if date_column < len(df_preview.columns):
            date_col_name = df_preview.columns[date_column]
            st.success(f"✅ Date Column [{date_column}]: `{date_col_name}`")
            extracted_preview['Date'] = df_preview.iloc[:, date_column]
        else:
            st.error(f"❌ Date column {date_column} doesn't exist!")

        # Show selected value columns
        for col_idx in new_selected:
            if col_idx < len(df_preview.columns):
                # Find column name
                try:
                    name_idx = available_cols.index(col_idx)
                    col_name = column_names[name_idx]
                except (ValueError, IndexError):
                    col_name = f"Column_{col_idx}"

                # Create final column name with file prefix
                final_name = f"{file_prefix} {col_name}"
                extracted_preview[final_name] = df_preview.iloc[:, col_idx]

        # Show timestamp conversion preview
        if date_column < len(df_preview.columns) and not extracted_preview.empty:
            st.markdown("**🕐 Timestamp Conversion:**")
            # Get first non-null timestamp
            sample_ts = extracted_preview['Date'].dropna().iloc[0] if len(extracted_preview['Date'].dropna()) > 0 else None
            if sample_ts is not None:
                try:
                    original_str = str(sample_ts)
                    normalized = format_timestamp_mdy_hms(original_str)

                    col_ts1, col_ts2, col_ts3 = st.columns([2, 1, 2])
                    with col_ts1:
                        st.text(f"Original: {original_str}")
                    with col_ts2:
                        st.text("→")
                    with col_ts3:
                        st.success(f"Standardized: {normalized}")
                except Exception as e:
                    st.caption(f"Timestamp will be normalized during combine step")

        # Show the extracted data table
        if not extracted_preview.empty:
            st.dataframe(prepare_df_for_display(extracted_preview), height=250)
        else:
            st.warning("No valid columns selected for extraction")
    else:
        st.error("Could not parse file with current settings")

    # Keep raw file preview in collapsible expander
    with st.expander("🔍 View Raw File Data (all columns)", expanded=False):
        if df_preview is not None and not df_preview.empty:
            st.caption("This shows ALL columns from the original file. Only the selected columns above will be used in processing.")
            st.dataframe(prepare_df_for_display(df_preview), height=300)


def render_stacked_config_ui(file_name, file_path, config):
    """
    Render configuration UI for stacked/long format files.

    Shows standard settings (start_row, delimiter, date_column), stacked-specific
    settings (time_column, equipment_column), detected equipment names preview,
    and value column checkboxes.
    """
    inner_config = config.get('config', config)

    st.info("**Stacked/Long Format Detected** - Data for multiple equipment groups shares the same timestamps. "
            "The data will be automatically pivoted to wide format (one column per equipment per value) during processing.")

    # Row 1: Basic settings
    col1, col2, col3 = st.columns(3)

    with col1:
        start_row = st.number_input(
            "Start Row", min_value=0, max_value=10,
            value=inner_config.get('start_row', 0),
            key=f"stacked_start_{file_name}",
            help="Row where column headers are located (0-based)"
        )

    with col2:
        delimiter_options = [',', '\t', ';', '|']
        delimiter_labels = ['Comma (,)', 'Tab (\\t)', 'Semicolon (;)', 'Pipe (|)']
        default_delim = inner_config.get('delimiter', ',')
        try:
            delim_idx = delimiter_options.index(default_delim)
        except ValueError:
            delim_idx = 0
        delimiter = st.selectbox(
            "Delimiter", options=delimiter_options,
            format_func=lambda x: delimiter_labels[delimiter_options.index(x)],
            index=delim_idx, key=f"stacked_delim_{file_name}"
        )

    with col3:
        date_column = st.number_input(
            "Date Column", min_value=0, max_value=20,
            value=inner_config.get('date_column', 0),
            key=f"stacked_date_{file_name}",
            help="Column index for date (0-based)"
        )

    # Row 2: Stacked-specific settings
    col4, col5 = st.columns(2)

    with col4:
        default_time = inner_config.get('time_column', None)
        time_column = st.number_input(
            "Time Column (or -1 if combined with date)",
            min_value=-1, max_value=20,
            value=default_time if default_time is not None else -1,
            key=f"stacked_time_{file_name}",
            help="Separate time column index, or -1 if timestamp is in a single column"
        )

    with col5:
        equipment_column = st.number_input(
            "Equipment/Identifier Column",
            min_value=0, max_value=20,
            value=inner_config.get('equipment_column', 1),
            key=f"stacked_equip_{file_name}",
            help="Column that identifies which equipment each row belongs to"
        )

    # Show detected equipment names (preview first 500 rows to find all groups)
    df_preview = parse_file_with_config(file_path, start_row, delimiter, num_rows=500)
    unique_equip = []
    if df_preview is not None and equipment_column < len(df_preview.columns):
        equip_series = df_preview.iloc[:, equipment_column].dropna().astype(str).str.strip()
        unique_equip = equip_series.unique().tolist()
        st.markdown(f"**Detected Equipment ({len(unique_equip)} groups):** {', '.join(unique_equip[:20])}")
        if len(unique_equip) > 20:
            st.caption(f"... and {len(unique_equip) - 20} more")

    # Value column checkboxes
    st.markdown("**Select Value Columns to Include (will be pivoted per equipment):**")

    available_cols = inner_config.get('available_columns', [])
    column_names = inner_config.get('column_names', [])
    currently_selected = inner_config.get('selected_columns', available_cols.copy())

    new_selected = []
    if available_cols:
        checkbox_cols = st.columns(3)
        for i, (col_idx, col_name) in enumerate(zip(available_cols, column_names)):
            is_checked = col_idx in currently_selected
            with checkbox_cols[i % 3]:
                if st.checkbox(
                    f"**{col_name}** (col {col_idx})",
                    value=is_checked,
                    key=f"stacked_col_{file_name}_{col_idx}"
                ):
                    new_selected.append(col_idx)

        n_equip = len(unique_equip) if unique_equip else 0
        if new_selected and n_equip > 0:
            st.success(f"{len(new_selected)} value column(s) x {n_equip} equipment groups = "
                       f"{len(new_selected) * n_equip} output columns")
        elif new_selected:
            st.success(f"{len(new_selected)} value column(s) selected")
        else:
            st.warning("No columns selected")
    else:
        # Fallback manual entry
        st.warning("No value columns detected by AI. Add manually:")
        manual_col = st.number_input("Value Column Index", min_value=0, max_value=20, value=3,
                                     key=f"stacked_manual_col_{file_name}")
        manual_name = st.text_input("Column Name", value="Value",
                                    key=f"stacked_manual_name_{file_name}")
        available_cols = [manual_col]
        column_names = [manual_name]
        new_selected = [manual_col]

    # Update config in session state
    effective_time_col = time_column if time_column >= 0 else None

    st.session_state.file_configs[file_name] = {
        'file_type': 'stacked_long',
        'config': {
            'start_row': start_row,
            'delimiter': delimiter,
            'date_column': date_column,
            'time_column': effective_time_col,
            'equipment_column': equipment_column,
            'is_stacked': True,
            'available_columns': available_cols,
            'column_names': column_names,
            'selected_columns': new_selected
        }
    }

    # Show preview of raw data
    with st.expander("View Raw File Data (first 10 rows)", expanded=False):
        if df_preview is not None:
            st.dataframe(prepare_df_for_display(df_preview.head(10)), height=300)


def extract_multi_tab_data(file_path, file_config):
    """
    Extract data from multi-tab Excel file.

    Args:
        file_path: Path to the Excel file
        file_config: Config dict with structure:
            {
                "file_type": "excel_multi_tab",
                "tabs": {
                    "Tab1": {
                        "start_row": 1,
                        "date_column": 0,
                        "available_columns": [2, 3, 4],
                        "column_names": ["Col1", "Col2", "Col3"],
                        "selected_columns": [2, 3]  # User-selected subset
                    }
                }
            }

    Returns:
        List of DataFrames, each with [Date, SensorName] columns
    """
    all_dataframes = []

    for tab_name, tab_config in file_config['tabs'].items():
        try:
            # Read the tab
            df = pd.read_excel(
                file_path,
                sheet_name=tab_name,
                header=tab_config['start_row'],
                dtype=str,
                keep_default_na=False
            )

            # Extract Date column
            date_col_idx = tab_config['date_column']
            date_series = df.iloc[:, date_col_idx]

            # Build dictionary with all selected columns for this tab
            selected_data = {}
            selected_data['Date'] = date_series

            # Extract selected value columns
            selected_indices = tab_config['selected_columns']
            column_names = tab_config['column_names']

            for col_idx in selected_indices:
                # Find position in available_columns to get correct name
                try:
                    name_idx = tab_config['available_columns'].index(col_idx)
                    col_name = column_names[name_idx]
                except (ValueError, IndexError):
                    col_name = f"Column_{col_idx}"

                # Create final column name: TabName ColumnName
                final_name = f"{tab_name} {col_name}"

                # Extract and intelligently convert value column (preserves text, converts numeric)
                value_series = smart_convert_column(df.iloc[:, col_idx], threshold=0.8)
                selected_data[final_name] = value_series

            # Create single DataFrame per tab with all selected columns
            tab_df = pd.DataFrame(selected_data)

            # Normalize timestamps ONCE per tab (not per column)
            tab_df['Date'] = tab_df['Date'].apply(
                lambda x: format_timestamp_mdy_hms(str(x)) if pd.notna(x) else None
            )

            # Convert Date to datetime
            tab_df['Date'] = pd.to_datetime(tab_df['Date'], format='%m/%d/%Y %H:%M:%S', errors='coerce')

            # Drop rows with invalid dates
            tab_df = tab_df.dropna(subset=['Date'])

            # Deduplicate by Date within each tab to prevent Cartesian products in outer join
            tab_df = tab_df.drop_duplicates(subset=['Date'], keep='first')

            if not tab_df.empty:
                all_dataframes.append(tab_df)

        except Exception as e:
            print(f"Error extracting data from tab {tab_name}: {str(e)}")
            continue

    return all_dataframes


def pivot_stacked_to_wide(df, config):
    """
    Transform stacked/long format data to wide format.

    Stacked format has multiple equipment groups sharing the same timestamps,
    listed sequentially. This function pivots the data so each equipment×value
    combination becomes its own column.

    Args:
        df: Raw DataFrame loaded from file (all string dtype)
        config: Inner config dict with keys:
            - date_column (int): column index for date
            - time_column (int or None): column index for separate time column
            - equipment_column (int): column index for equipment/identifier
            - selected_columns (list[int]): value column indices
            - available_columns (list[int]): all detected value column indices
            - column_names (list[str]): names for available_columns

    Returns:
        Wide-format DataFrame with columns: [Date, "equip value1", "equip value2", ...]
        Date column is datetime type, value columns are smart-converted.
        Returns None if pivot fails.
    """
    try:
        # 1. MERGE DATE + TIME if split
        date_col_idx = config['date_column']
        time_col_idx = config.get('time_column', None)

        date_series = df.iloc[:, date_col_idx].astype(str)

        if time_col_idx is not None and time_col_idx >= 0:
            time_series = df.iloc[:, time_col_idx].astype(str)
            merged_ts = date_series.str.strip() + ' ' + time_series.str.strip()
        else:
            merged_ts = date_series

        # 2. NORMALIZE TIMESTAMPS
        normalized_dates = []
        for ts_val in merged_ts:
            try:
                normalized = format_timestamp_mdy_hms(str(ts_val))
                normalized_dates.append(pd.to_datetime(normalized, format='%m/%d/%Y %H:%M:%S'))
            except Exception:
                try:
                    normalized_dates.append(pd.to_datetime(ts_val, format='mixed', errors='coerce'))
                except Exception:
                    normalized_dates.append(pd.NaT)

        df = df.copy()
        df['__Date__'] = normalized_dates

        # Drop rows with invalid dates
        df = df.dropna(subset=['__Date__'])

        # Remove timezone if present
        if pd.api.types.is_datetime64tz_dtype(df['__Date__']):
            df['__Date__'] = df['__Date__'].dt.tz_localize(None)

        # 3. GET EQUIPMENT COLUMN
        equip_col_idx = config['equipment_column']
        df['__Equipment__'] = df.iloc[:, equip_col_idx].astype(str).str.strip()
        df = df.dropna(subset=['__Equipment__'])

        # 4. EXTRACT SELECTED VALUE COLUMNS
        selected_cols = config.get('selected_columns', [])
        available_cols = config.get('available_columns', [])
        column_names = config.get('column_names', [])

        value_col_names = []
        for col_idx in selected_cols:
            if col_idx < len(df.columns):
                try:
                    name_idx = available_cols.index(col_idx)
                    col_name = column_names[name_idx]
                except (ValueError, IndexError):
                    col_name = f"Column_{col_idx}"
                value_col_names.append(col_name)

        # 5. BUILD PIVOT-READY DATAFRAME
        pivot_data = {'Date': df['__Date__'], 'Equipment': df['__Equipment__']}
        for col_idx, col_name in zip(selected_cols, value_col_names):
            if col_idx < len(df.columns):
                pivot_data[col_name] = smart_convert_column(df.iloc[:, col_idx], threshold=0.8).values

        pivot_df = pd.DataFrame(pivot_data)

        # 6. PIVOT: long -> wide
        wide_df = pivot_df.pivot_table(
            index='Date',
            columns='Equipment',
            values=value_col_names,
            aggfunc='first'
        )

        # 7. FLATTEN MULTI-LEVEL COLUMN INDEX
        # pivot_table creates MultiIndex: (value_name, equipment_name)
        # Flatten to: "equipment_name value_name"
        wide_df.columns = [f"{equip} {val}" for val, equip in wide_df.columns]

        # 8. RESET INDEX to make Date a regular column
        wide_df = wide_df.reset_index()

        # 9. SORT columns alphabetically (equipment grouped)
        date_col = ['Date']
        sensor_cols = sorted([c for c in wide_df.columns if c != 'Date'])
        wide_df = wide_df[date_col + sensor_cols]

        return wide_df

    except Exception as e:
        print(f"Error in pivot_stacked_to_wide: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def detect_percentage_columns(file_path, sheet_name, col_indices):
    """
    Check if columns have percentage formatting in source Excel file.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet/tab
        col_indices: List of column indices to check

    Returns:
        List of column indices that have percentage formatting
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=False)

        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        percentage_cols = []

        for col_idx in col_indices:
            # Check first data row (row 2 assuming row 1 is header)
            cell = ws.cell(row=2, column=col_idx + 1)  # Excel is 1-indexed
            if cell.number_format and '%' in cell.number_format:
                percentage_cols.append(col_idx)

        wb.close()
        return percentage_cols

    except Exception as e:
        print(f"Error detecting percentage columns: {str(e)}")
        return []


def export_to_excel(resampled_df, inexact_df, output_path):
    """
    Export resampled data to Excel with color-coded quality indicators.

    V12 Optimized: Uses Bulk Write + Sparse Styling for performance.
    - Bulk writes data using pandas (fast C code instead of Python loops)
    - Only visits cells that need coloring (sparse iteration with np.where)

    Features:
    - Inexact cells are highlighted in yellow
    - Stale_Data_Flag column highlights True values in light red
    - Zero_Value_Flag column included (Clear/Single/Repeated)
    - Date column formatted as text to preserve formatting

    Args:
        resampled_df: The resampled DataFrame
        inexact_df: Boolean DataFrame with sensors as columns, rows matching resampled_df
        output_path: Path to save the Excel file

    Returns:
        True if successful, False otherwise
    """
    try:
        # 1. PREPARE DATA
        export_df = resampled_df.copy()

        # Format Date as string to preserve formatting
        if 'Date' in export_df.columns:
            export_df['Date'] = export_df['Date'].dt.strftime('%m/%d/%Y %H:%M:%S')

        # 2. BULK WRITE (fast - uses optimized C code)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Resampled Data')

            wb = writer.book
            ws = writer.sheets['Resampled Data']

            # Define Styles
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # 3. SPARSE STYLING (smart - only visit cells that need color)
            col_map = {name: i + 1 for i, name in enumerate(export_df.columns)}

            # A. Color Stale Flags (Red)
            if 'Stale_Data_Flag' in export_df.columns:
                stale_col_idx = col_map['Stale_Data_Flag']
                # Use numpy to find True values instantly
                stale_mask = export_df['Stale_Data_Flag'].values == True
                stale_indices = np.where(stale_mask)[0]

                for r_idx in stale_indices:
                    # Excel row = index + 2 (1 for header, 1 because index starts at 0)
                    ws.cell(row=r_idx + 2, column=stale_col_idx).fill = red_fill

            # B. Color Inexact Cells (Yellow)
            safe_inexact_df = inexact_df.fillna(False)
            common_cols = [c for c in safe_inexact_df.columns if c in col_map]

            for col_name in common_cols:
                excel_col_idx = col_map[col_name]
                # Find only the "bad" cells for this column
                bad_indices = np.where(safe_inexact_df[col_name].values)[0]

                for r_idx in bad_indices:
                    ws.cell(row=r_idx + 2, column=excel_col_idx).fill = yellow_fill

            # 4. SAMPLE-BASED AUTO-WIDTH (fast - only check first 10 rows)
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                # Check header + first 10 data rows only
                max_len = len(str(col_cells[0].value or ''))
                for cell in col_cells[1:11]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        return True

    except Exception as e:
        print(f"Export error: {str(e)}")
        # Fallback: save as CSV if Excel fails
        try:
            resampled_df.to_csv(str(output_path).replace('.xlsx', '.csv'), index=False)
        except:
            pass
        return False


def archive_uploaded_files(uploaded_files, archive_path):
    """
    Copy original files to archive directory for safekeeping.

    Args:
        uploaded_files: Dictionary of {filename: filepath}
        archive_path: Destination directory path

    Returns:
        List of archived file paths
    """
    try:
        # Create archive directory
        archive_dir = Path(archive_path)
        archive_dir.mkdir(parents=True, exist_ok=True)

        # Copy each file
        archived_files = []
        for file_name, file_path in uploaded_files.items():
            dest_path = archive_dir / file_name
            shutil.copy2(file_path, dest_path)
            archived_files.append(str(dest_path))

        return archived_files

    except Exception as e:
        raise Exception(f"Archiving failed: {str(e)}")


def sanitize_building_name(building_name):
    """
    Remove special characters from building name for use in filenames.

    Args:
        building_name: Building name string to sanitize

    Returns:
        Sanitized string safe for filenames
    """
    if not building_name or not building_name.strip():
        return "Unnamed"

    # Remove special characters, keep alphanumeric, spaces, hyphens, underscores
    safe_name = re.sub(r'[^\w\s-]', '', building_name)
    # Replace spaces with underscores
    safe_name = safe_name.strip().replace(' ', '_')
    # Remove consecutive underscores
    safe_name = re.sub(r'_+', '_', safe_name)

    return safe_name


def validate_archive_path(archive_path):
    """
    Validate archive path exists and is writable.

    Args:
        archive_path: Path to validate

    Returns:
        Tuple of (is_valid: bool, error_message: str or None)
    """
    path = Path(archive_path)

    # Check if path exists, create if not
    if not path.exists():
        try:
            path.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return False, f"Cannot create archive path: {e}"

    # Check if writable
    if not os.access(str(path), os.W_OK):
        return False, f"Archive path is not writable: {archive_path}"

    return True, None


def auto_process_and_export(
    file_configs,
    uploaded_files,
    archive_path,
    building_name,
    progress_callback=None
):
    """
    Orchestrate entire automatic workflow: combine -> save CSV -> resample -> save Excel.

    V12: Now handles multi-column CSV files using the same extraction pattern as multi-tab Excel.

    Args:
        file_configs: Dictionary of file configurations
        uploaded_files: Dictionary of {filename: filepath}
        archive_path: Archive folder path
        building_name: Building name for file naming
        progress_callback: Optional callback(phase, current, total, message)

    Returns:
        Tuple of (success: bool, results: dict)
        Results dict contains: combined_df, resampled_df, raw_csv_path, excel_path, stats, inexact_cells
        On error: results dict contains: error (and partial results if available)
    """
    try:
        # Validate archive path first
        valid, error = validate_archive_path(archive_path)
        if not valid:
            return False, {'error': error}

        # ===== PHASE 1: COMBINE DATA (0-40%) =====
        if progress_callback:
            progress_callback('combine', 0, len(file_configs), "Starting data combination...")

        loaded_dfs = []
        total_files = len(file_configs)

        for idx, (file_name, config) in enumerate(file_configs.items()):
            if progress_callback:
                progress_callback('combine', idx + 1, total_files,
                                f"Processing {file_name}... ({idx + 1}/{total_files})")

            file_path = uploaded_files[file_name]
            file_type = config.get('file_type', 'csv')

            # Multi-tab Excel file
            if file_type == 'excel_multi_tab':
                multi_tab_dfs = extract_multi_tab_data(file_path, config)
                loaded_dfs.extend(multi_tab_dfs)

            # Stacked/long format file - pivot to wide before adding
            elif file_type == 'stacked_long':
                inner_config = config.get('config', config)

                # Read full file
                if str(file_path).lower().endswith(('.xlsx', '.xls')):
                    df_full = pd.read_excel(
                        file_path,
                        header=inner_config['start_row'],
                        dtype=str,
                        keep_default_na=False
                    )
                else:
                    df_full = pd.read_csv(
                        file_path,
                        sep=inner_config['delimiter'],
                        header=inner_config['start_row'],
                        dtype=str,
                        keep_default_na=False,
                        encoding='utf-8',
                        encoding_errors='ignore',
                        on_bad_lines='skip'
                    )

                # Pivot stacked data to wide format
                wide_df = pivot_stacked_to_wide(df_full, inner_config)

                if wide_df is not None and not wide_df.empty:
                    # Remove timezone if present
                    if pd.api.types.is_datetime64tz_dtype(wide_df['Date']):
                        wide_df['Date'] = wide_df['Date'].dt.tz_localize(None)

                    # Deduplicate by Date (safe now because data is wide after pivot)
                    wide_df = wide_df.drop_duplicates(subset=['Date'], keep='first')

                    loaded_dfs.append(wide_df)
                else:
                    print(f"Warning: pivot_stacked_to_wide returned empty for {file_name}")

            # CSV or single-tab Excel file (V12: now supports multi-column)
            else:
                inner_config = config.get('config', config)

                # Read full file
                if str(file_path).lower().endswith(('.xlsx', '.xls')):
                    df_full = pd.read_excel(
                        file_path,
                        header=inner_config['start_row'],
                        dtype=str,
                        keep_default_na=False
                    )
                else:
                    df_full = pd.read_csv(
                        file_path,
                        sep=inner_config['delimiter'],
                        header=inner_config['start_row'],
                        dtype=str,
                        keep_default_na=False,
                        encoding='utf-8',
                        encoding_errors='ignore',
                        on_bad_lines='skip'
                    )

                # Extract required columns
                df_clean = pd.DataFrame()

                # Get file prefix for column naming
                file_prefix = Path(file_name).stem

                # Get date column
                if inner_config['date_column'] < len(df_full.columns):
                    date_col_name = df_full.columns[inner_config['date_column']]
                    df_clean['Date'] = df_full[date_col_name]

                # V12: Handle multi-column extraction (loop through selected_columns)
                selected_cols = inner_config.get('selected_columns', [])
                available_cols = inner_config.get('available_columns', [])
                column_names = inner_config.get('column_names', [])

                # Backward compatibility: check for old single-column format
                if not selected_cols and 'value_column' in inner_config:
                    # Old format - single column
                    selected_cols = [inner_config['value_column']]
                    available_cols = [inner_config['value_column']]
                    column_names = [inner_config.get('sensor_name', f"Column_{inner_config['value_column']}")]

                for col_idx in selected_cols:
                    if col_idx < len(df_full.columns):
                        # Find column name
                        try:
                            name_idx = available_cols.index(col_idx)
                            col_name = column_names[name_idx]
                        except (ValueError, IndexError):
                            col_name = f"Column_{col_idx}"

                        # Create final column name with file prefix: "Filename ColumnName"
                        final_name = f"{file_prefix} {col_name}"

                        # Extract with smart conversion
                        df_clean[final_name] = smart_convert_column(
                            df_full.iloc[:, col_idx], threshold=0.8
                        )

                # Normalize timestamps
                if 'Date' in df_clean.columns:
                    normalized_dates = []
                    for ts_val in df_clean['Date']:
                        try:
                            normalized = format_timestamp_mdy_hms(str(ts_val))
                            normalized_dates.append(pd.to_datetime(normalized, format='%m/%d/%Y %H:%M:%S'))
                        except Exception:
                            try:
                                normalized_dates.append(pd.to_datetime(ts_val, format='mixed', errors='coerce'))
                            except:
                                normalized_dates.append(pd.NaT)

                    df_clean['Date'] = normalized_dates
                    df_clean = df_clean.dropna(subset=['Date'])

                    # Remove timezone if present
                    if pd.api.types.is_datetime64tz_dtype(df_clean['Date']):
                        df_clean['Date'] = df_clean['Date'].dt.tz_localize(None)

                    # Deduplicate by Date within each file to prevent Cartesian products in outer join
                    df_clean = df_clean.drop_duplicates(subset=['Date'], keep='first')

                    loaded_dfs.append(df_clean)

        # Merge all DataFrames
        if not loaded_dfs:
            return False, {'error': 'No data frames loaded from files'}

        from functools import reduce
        combined = reduce(
            lambda left, right: pd.merge(left, right, on='Date', how='outer'),
            loaded_dfs
        )

        # Sort and deduplicate
        combined = combined.sort_values('Date').reset_index(drop=True)
        combined = combined.drop_duplicates()

        # ===== PHASE 2: SAVE RAW CSV (40%) =====
        if progress_callback:
            progress_callback('export', 1, 3, "Saving raw merged CSV...")

        # Generate filename
        safe_name = sanitize_building_name(building_name)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        raw_csv_filename = f"{safe_name}_raw_merged_{timestamp}.csv"

        # Save to archive folder
        archive_dir = Path(archive_path)
        raw_csv_path = archive_dir / raw_csv_filename

        # Export with formatted dates
        combined_export = combined.copy()
        combined_export['Date'] = combined_export['Date'].dt.strftime('%m/%d/%Y %H:%M:%S')
        combined_export.to_csv(raw_csv_path, index=False)

        # Verify file exists
        if not raw_csv_path.exists():
            return False, {
                'error': 'Raw CSV file was not created',
                'combined_df': combined
            }

        # ===== PHASE 3: RESAMPLE TO 15-MIN (40-80%) =====
        if progress_callback:
            progress_callback('resample', 0, 1, "Starting quarter-hour resampling...")

        def resample_progress(current, total, message):
            if progress_callback:
                progress_callback('resample', current, total, message)

        resampled_df, stats, inexact_cells = resample_to_quarter_hour(
            combined,
            tolerance_minutes=2,
            progress_callback=resample_progress
        )

        if resampled_df is None:
            return False, {
                'error': 'Resampling failed - returned None',
                'combined_df': combined,
                'raw_csv_path': str(raw_csv_path)
            }

        # ===== PHASE 4: GENERATE EXCEL (80-100%) =====
        if progress_callback:
            progress_callback('export', 2, 3, "Generating Excel file with color coding...")

        # Generate Excel filename
        excel_filename = f"{safe_name}_resampled_15min_{timestamp}.xlsx"
        excel_path = archive_dir / excel_filename

        # Export to Excel with color coding
        success = export_to_excel(resampled_df, inexact_cells, excel_path)

        if not success or not excel_path.exists():
            return False, {
                'error': 'Excel file was not created',
                'combined_df': combined,
                'resampled_df': resampled_df,
                'raw_csv_path': str(raw_csv_path),
                'stats': stats
            }

        if progress_callback:
            progress_callback('export', 3, 3, "Export complete!")

        # Return success with all results
        return True, {
            'combined_df': combined,
            'resampled_df': resampled_df,
            'raw_csv_path': str(raw_csv_path),
            'excel_path': str(excel_path),
            'stats': stats,
            'inexact_cells': inexact_cells
        }

    except Exception as e:
        return False, {
            'error': f'Processing error: {str(e)}',
            # Include partial results if available
            'combined_df': combined if 'combined' in locals() else None,
            'resampled_df': resampled_df if 'resampled_df' in locals() else None,
            'raw_csv_path': str(raw_csv_path) if 'raw_csv_path' in locals() and raw_csv_path.exists() else None
        }


def main():
    """Main application interface."""

    # Subheader below logo
    st.markdown("**Multi-Column CSV** • **Automatic Workflow** • **Tab-Based UI** • **Smart Data Types** • **4-Step Process**")
    st.markdown("---")

    # ========== STEP 1: UPLOAD ==========
    st.header("Step 1: Upload Files & Archive Settings")

    # Building Name Input (Required for archiving)
    building_name = st.text_input(
        "Building Name",
        value=st.session_state.get('building_name', ''),
        key="building_name",
        placeholder="e.g., Gotham Tower",
        help="Used for organizing archived files"
    )

    # Archive Location Selection
    st.markdown("### 📦 Archive Settings")

    use_custom_path = st.checkbox(
        "Use custom archive location",
        value=st.session_state.get('use_custom_archive', False),
        key="use_custom_archive",
        help="Check this to manually specify where to save archived files"
    )

    if use_custom_path:
        # Manual filepath entry
        archive_path = st.text_input(
            "Custom Archive Folder Path",
            value=st.session_state.get('archive_path', ''),
            key="archive_path",
            placeholder="e.g., C:/My Files/Archive or D:/Backups",
            help="Enter the full path where you want to save the original files"
        )
        st.caption("💡 Enter a custom path above (e.g., `C:/MyFolder/Archive` or `D:/Backups/BuildingData`)")
    else:
        # Default archive structure: archive/[Building Name]/
        if building_name and building_name.strip():
            default_archive = f"archive/{building_name}"
        else:
            default_archive = "archive/Unnamed"

        archive_path = default_archive
        st.session_state.archive_path = default_archive

        st.info(f"📁 **Default Archive Location:** `{default_archive}/`")
        st.caption("Files will be saved to the archive folder in your project directory")

    st.caption("📦 **All uploaded files will be archived** before processing")
    st.markdown("")  # Spacing

    # File Uploader
    uploaded_files = st.file_uploader(
        "📁 Upload your sensor files",
        type=['xlsx', 'csv', 'xls'],
        accept_multiple_files=True,
        help="Select all files (10-90 files)"
    )

    if uploaded_files:
        # Save files to temp directory
        temp_dir = Path("temp")
        temp_dir.mkdir(exist_ok=True)

        for uploaded_file in uploaded_files:
            file_path = temp_dir / uploaded_file.name
            if uploaded_file.name not in st.session_state.uploaded_files:
                with open(file_path, "wb") as f:
                    f.write(uploaded_file.getbuffer())
                st.session_state.uploaded_files[uploaded_file.name] = str(file_path)

        # Always archive files to the specified path
        archive_path = st.session_state.get('archive_path', '')

        if archive_path and archive_path.strip():
            try:
                archived_files = archive_uploaded_files(
                    st.session_state.uploaded_files,
                    archive_path
                )
                st.success(f"✅ {len(uploaded_files)} files uploaded and archived to:")
                st.code(archive_path, language="")
            except Exception as e:
                st.error(f"❌ Archiving failed: {e}")
                st.warning("Files are saved in temp storage but not permanently archived.")
                st.info("Please check the archive path and try re-uploading, or proceed with processing anyway.")
        else:
            st.warning("⚠️ Archive path is empty - files uploaded to temp storage only")
            st.info("Files will be processed but not permanently archived")

        st.markdown("---")

        # ========== STEP 2: AI ANALYSIS ==========
        st.header("Step 2: AI Analysis")

        # Check for API key
        api_key = os.getenv('CLAUDE_API_KEY')
        if not api_key:
            st.error("⚠️ CLAUDE_API_KEY not found in .env file. Please add your API key.")
            st.code("CLAUDE_API_KEY=sk-ant-...", language="bash")
        else:
            col1, col2 = st.columns([2, 1])

            with col1:
                st.info("💡 Click below to analyze ALL files in parallel with Claude AI")

            with col2:
                if st.button("🤖 Analyze All Files", type="primary"):
                    with st.spinner(f"Analyzing {len(st.session_state.uploaded_files)} files in parallel..."):
                        progress_bar = st.progress(0)

                        # Run parallel analysis
                        configs, debug_logs = analyze_all_files_parallel(
                            st.session_state.uploaded_files,
                            api_key
                        )

                        progress_bar.progress(100)

                        # Store results
                        st.session_state.file_configs = configs
                        st.session_state.ai_debug_log = debug_logs
                        st.session_state.ai_analysis_complete = True

                        # Show summary
                        success_count = sum(1 for log in debug_logs if log.get('success', False))
                        st.success(f"✅ Analysis complete! {success_count}/{len(debug_logs)} files configured successfully")
                        st.rerun()

            # Show analysis results if complete
            if st.session_state.ai_analysis_complete and st.session_state.file_configs:
                st.markdown("---")
                st.subheader("Analysis Results")

                # Summary metrics
                total_files = len(st.session_state.uploaded_files)
                configured_files = len(st.session_state.file_configs)
                success_rate = (configured_files / total_files * 100) if total_files > 0 else 0

                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Files", total_files)
                with col2:
                    st.metric("Configured", configured_files)
                with col3:
                    st.metric("Success Rate", f"{success_rate:.0f}%")

        # ========== STEP 3: REVIEW & EDIT CONFIGS ==========
        if st.session_state.file_configs:
            st.markdown("---")
            st.header("Step 3: Review & Edit Configurations")

            st.info("💡 Review AI-detected settings. Click tabs to configure each file. **V12: CSV files now support multiple value columns!**")

            # Build file-level tabs
            file_tab_labels = []
            file_names_ordered = list(st.session_state.uploaded_files.keys())

            for file_name in file_names_ordered:
                config = st.session_state.file_configs.get(file_name, {})
                file_type = config.get('file_type', 'csv')

                # Calculate total selected columns
                if file_type == 'excel_multi_tab':
                    total_selected = sum(
                        len(tab_cfg.get('selected_columns', []))
                        for tab_cfg in config['tabs'].values()
                    )
                    total_available = sum(
                        len(tab_cfg.get('available_columns', []))
                        for tab_cfg in config['tabs'].values()
                    )
                else:
                    # V12: CSV now uses arrays too
                    inner_config = config.get('config', {})
                    total_selected = len(inner_config.get('selected_columns', []))
                    total_available = len(inner_config.get('available_columns', []))

                label = build_tab_label(file_name, total_selected, total_available)
                file_tab_labels.append(label)

            # Render file-level tabs
            file_tabs = st.tabs(file_tab_labels)

            for file_tab, file_name in zip(file_tabs, file_names_ordered):
                with file_tab:
                    config = st.session_state.file_configs.get(file_name, {})
                    file_type = config.get('file_type', 'csv')
                    file_path = st.session_state.uploaded_files[file_name]

                    if file_type == 'excel_multi_tab':
                        st.info(f"📑 **Multi-Tab Excel** - {len(config['tabs'])} sheets")

                        # Build sheet-level tabs
                        sheet_tab_labels = []
                        sheet_names_ordered = list(config['tabs'].keys())

                        for sheet_name in sheet_names_ordered:
                            tab_cfg = config['tabs'][sheet_name]
                            selected_count = len(tab_cfg.get('selected_columns', []))
                            available_count = len(tab_cfg.get('available_columns', []))
                            label = build_tab_label(sheet_name, selected_count, available_count)
                            sheet_tab_labels.append(label)

                        # Render sheet-level tabs
                        sheet_tabs = st.tabs(sheet_tab_labels)

                        for sheet_tab, sheet_name in zip(sheet_tabs, sheet_names_ordered):
                            with sheet_tab:
                                render_sheet_config_ui(file_name, file_path, sheet_name, config)

                        # Update config
                        st.session_state.file_configs[file_name] = config

                    elif file_type == 'stacked_long':
                        # Stacked/long format: show pivot-specific config UI
                        render_stacked_config_ui(file_name, file_path, config)

                    else:
                        # CSV/Single-tab: Render config directly (V12: now with multi-column support)
                        render_csv_config_ui(file_name, file_path, config)

        # ========== STEP 4: PROCESS & EXPORT (AUTOMATIC) ==========
        if st.session_state.file_configs and len(st.session_state.file_configs) == len(st.session_state.uploaded_files):
            st.markdown("---")
            st.header("Step 4: Process & Export All Data")

            # Check if processing already complete
            if not st.session_state.processing_complete:
                # BEFORE PROCESSING: Show info and button

                st.info(f"""
                🚀 **Automatic Workflow:**
                1. **Combine** all sensor data (outer join on timestamps)
                2. **Normalize** timestamps to MM/DD/YYYY HH:MM:SS format
                3. **Save raw CSV** to archive folder
                4. **Resample** to 15-minute intervals (per-sensor matching +/-2 min)
                5. **Apply quality flags** (stale data + zero value tracking)
                6. **Generate Excel** with color-coded quality indicators
                7. **Provide downloads** for both files

                **All files will be saved to:** `{st.session_state.archive_path}`
                """)

                # Timestamp conversion preview (keep existing expander)
                with st.expander("📅 Preview Timestamp Conversion", expanded=False):
                    st.markdown("**Sample conversions from your uploaded files:**")

                    for file_name, config in list(st.session_state.file_configs.items())[:3]:
                        file_type = config.get('file_type', 'csv')

                        if file_type == 'excel_multi_tab':
                            st.info(f"**{file_name}**: Multi-tab Excel - timestamps normalized during processing")
                            continue

                        inner_config = config.get('config', config)
                        file_path = st.session_state.uploaded_files[file_name]
                        df_sample = parse_file_with_config(
                            file_path,
                            inner_config.get('start_row', 0),
                            inner_config.get('delimiter', ','),
                            num_rows=3
                        )

                        if df_sample is not None and inner_config.get('date_column', 0) < len(df_sample.columns):
                            st.markdown(f"**{file_name}:**")
                            date_col_name = df_sample.columns[inner_config['date_column']]
                            sample_timestamps = df_sample[date_col_name].dropna().head(2)

                            # For stacked files with split date/time, merge columns for preview
                            time_col_idx = inner_config.get('time_column', None)
                            sample_times = None
                            if time_col_idx is not None and time_col_idx >= 0 and time_col_idx < len(df_sample.columns):
                                time_col_name = df_sample.columns[time_col_idx]
                                sample_times = df_sample[time_col_name].dropna().head(2)

                            for i, original_ts in enumerate(sample_timestamps):
                                try:
                                    original_str = str(original_ts).strip()
                                    # Merge date + time if split columns
                                    if sample_times is not None and i < len(sample_times):
                                        original_str = original_str + ' ' + str(sample_times.iloc[i]).strip()
                                    detected_format = detect_timestamp_format(original_str)
                                    normalized = format_timestamp_mdy_hms(original_str)

                                    col1, col2, col3 = st.columns([2, 1, 2])
                                    with col1:
                                        st.text(f"Original: {original_str}")
                                    with col2:
                                        st.text("→")
                                    with col3:
                                        st.success(f"Standardized: {normalized}")
                                    st.caption(f"Detected: {detected_format}")
                                    st.markdown("---")
                                except Exception as e:
                                    st.warning(f"Could not parse: {original_ts}")

                # Single button to trigger entire workflow
                if st.button("🚀 Process All Files", type="primary", key="process_all_btn"):
                    # Create progress tracking UI
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    # Progress callback function
                    def update_progress(phase, current, total, message):
                        """Calculate and update progress bar."""
                        # Phase weights: combine=40%, resample=40%, export=20%
                        phase_weights = {
                            'combine': (0.0, 0.4),
                            'resample': (0.4, 0.8),
                            'export': (0.8, 1.0)
                        }

                        start, end = phase_weights.get(phase, (0.0, 1.0))
                        phase_progress = current / total if total > 0 else 0
                        overall = start + (end - start) * phase_progress

                        progress_bar.progress(min(overall, 1.0))
                        status_text.text(message)

                    try:
                        # Call automatic processing function
                        success, results = auto_process_and_export(
                            st.session_state.file_configs,
                            st.session_state.uploaded_files,
                            st.session_state.archive_path,
                            st.session_state.building_name,
                            progress_callback=update_progress
                        )

                        if success:
                            # Store results in session state
                            st.session_state.combined_df = results['combined_df']
                            st.session_state.resampled_df = results['resampled_df']
                            st.session_state.resampling_stats = results['stats']
                            st.session_state.inexact_cells = results['inexact_cells']
                            st.session_state.raw_csv_path = results['raw_csv_path']
                            st.session_state.excel_output_path = results['excel_path']
                            st.session_state.processing_complete = True

                            # Clear progress indicators
                            progress_bar.empty()
                            status_text.empty()

                            st.success("✅ Processing complete! Files saved to archive folder.")
                            st.balloons()
                            st.rerun()
                        else:
                            # Show error
                            progress_bar.empty()
                            status_text.empty()
                            st.error(f"❌ Processing failed: {results.get('error', 'Unknown error')}")

                    except Exception as e:
                        progress_bar.empty()
                        status_text.empty()
                        st.error(f"❌ Unexpected error: {str(e)}")
                        st.exception(e)

            else:
                # AFTER PROCESSING: Show results and downloads

                st.success("✅ Processing complete! Files are ready for download.")

                # Summary metrics
                st.markdown("### Processing Summary")

                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    combined_rows = len(st.session_state.combined_df)
                    st.metric("Combined Rows", f"{combined_rows:,}")
                with col2:
                    total_rows = len(st.session_state.resampled_df)
                    st.metric("Resampled Intervals", f"{total_rows:,}")
                with col3:
                    num_sensors = len([c for c in st.session_state.resampled_df.columns
                                     if c not in ['Date', 'Stale_Data_Flag', 'Stale_Sensors', 'Zero_Value_Flag']])
                    st.metric("Sensors", num_sensors)
                with col4:
                    df = st.session_state.resampled_df
                    date_range = f"{df['Date'].min().strftime('%m/%d/%Y')} - {df['Date'].max().strftime('%m/%d/%Y')}"
                    st.metric("Date Range", date_range)
                with col5:
                    stale_rows = st.session_state.resampling_stats.get('rows_with_stale_data', 0)
                    st.metric("Rows w/ Stale Data", stale_rows)

                # Download buttons
                st.markdown("---")
                st.markdown("### 📥 Download Files")

                col_left, col_right = st.columns(2)

                with col_left:
                    st.markdown("#### Raw Merged CSV")
                    st.caption("All sensor data combined with original timestamps")

                    if st.session_state.raw_csv_path and Path(st.session_state.raw_csv_path).exists():
                        with open(st.session_state.raw_csv_path, 'rb') as f:
                            csv_data = f.read()

                        filename = Path(st.session_state.raw_csv_path).name
                        st.download_button(
                            label="⬇️ Download Raw CSV",
                            data=csv_data,
                            file_name=filename,
                            mime='text/csv',
                            key="dl_raw_csv"
                        )
                        st.info(f"📁 Saved to: `{st.session_state.raw_csv_path}`")
                    else:
                        st.error("Raw CSV file not found")

                with col_right:
                    st.markdown("#### Resampled 15-Min Excel")
                    st.caption("Quarter-hour intervals with color-coded quality flags")

                    if st.session_state.excel_output_path and Path(st.session_state.excel_output_path).exists():
                        with open(st.session_state.excel_output_path, 'rb') as f:
                            excel_data = f.read()

                        filename = Path(st.session_state.excel_output_path).name
                        st.download_button(
                            label="⬇️ Download Excel",
                            data=excel_data,
                            file_name=filename,
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key="dl_excel"
                        )
                        st.info(f"📁 Saved to: `{st.session_state.excel_output_path}`")
                    else:
                        st.error("Excel file not found")

                # Data preview tabs
                st.markdown("---")
                st.markdown("### 📊 Data Preview")

                tab1, tab2 = st.tabs(["Resampled Data (15-min)", "Quality Statistics"])

                with tab1:
                    st.dataframe(
                        prepare_df_for_display(st.session_state.resampled_df.head(50)),
                        height=400
                    )

                with tab2:
                    stats = st.session_state.resampling_stats

                    col1, col2 = st.columns(2)

                    with col1:
                        st.markdown("#### Quality Flags Summary")
                        st.write(f"**Total Intervals**: {stats.get('total_intervals', 0):,}")
                        st.write(f"**Inexact Cells** (yellow): {stats.get('total_inexact_cells', 0):,}")
                        st.write(f"**Rows with Stale Data** (red): {stats.get('rows_with_stale_data', 0):,}")
                        st.write(f"**Total Stale Flags**: {stats.get('total_stale_flags', 0):,}")

                        st.markdown("#### Zero Value Flags")
                        zero_counts = stats.get('zero_flag_counts', {})
                        st.write(f"**Clear**: {zero_counts.get('Clear', 0):,}")
                        st.write(f"**Single**: {zero_counts.get('Single', 0):,}")
                        st.write(f"**Repeated**: {zero_counts.get('Repeated', 0):,}")

                    with col2:
                        st.markdown("#### Stale Data by Sensor")
                        stale_by_sensor = stats.get('stale_by_sensor', {})
                        if stale_by_sensor:
                            stale_df = pd.DataFrame([
                                {'Sensor': sensor, 'Stale Count': count}
                                for sensor, count in stale_by_sensor.items()
                                if count > 0
                            ])

                            if not stale_df.empty:
                                stale_df = stale_df.sort_values('Stale Count', ascending=False)
                                st.dataframe(stale_df, height=300)
                            else:
                                st.success("✅ No stale data detected!")
                        else:
                            st.info("No stale data information available")

                # Reset button
                st.markdown("---")
                if st.button("🔄 Process Different Files", type="secondary"):
                    # Clear processing state
                    st.session_state.combined_df = None
                    st.session_state.resampled_df = None
                    st.session_state.resampling_stats = {}
                    st.session_state.inexact_cells = pd.DataFrame()  # Reset to empty DataFrame
                    st.session_state.raw_csv_path = None
                    st.session_state.excel_output_path = None
                    st.session_state.processing_complete = False
                    st.rerun()


if __name__ == "__main__":
    main()
